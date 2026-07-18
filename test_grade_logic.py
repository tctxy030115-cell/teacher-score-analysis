from io import BytesIO
import unittest
from unittest.mock import patch

import pandas as pd
from openpyxl import load_workbook

import grade_logic

from grade_logic import (
    CLASS_COLUMN_ALIASES,
    NAME_COLUMN_ALIASES,
    analyze_scores,
    build_dataframe_from_header,
    clean_column_name,
    create_single_score_template,
    detect_header_row,
    export_score_result_to_bytes,
    find_first_matching_column,
    has_analyzable_columns,
    format_class_value,
    build_class_options,
    get_score_level,
)


class GradeLogicTest(unittest.TestCase):
    def test_normalize_score_column_name_supports_common_suffixes_and_total_aliases(self):
        self.assertTrue(
            hasattr(grade_logic, "normalize_score_column_name"),
            "缺少 normalize_score_column_name",
        )

        cases = {
            " 数学分数\n": "数学",
            "英语成绩": "英语",
            "物理得分": "物理",
            "体育分": "体育",
            "总分分数": "总分",
            "总成绩": "总分",
            "总成绩分数": "总分",
            "合计": "总分",
            "合计分数": "总分",
            "班级分数线": "班级分数线",
            "年级平均分": "年级平均",
        }
        for original, expected in cases.items():
            with self.subTest(original=original):
                self.assertEqual(grade_logic.normalize_score_column_name(original), expected)

    def test_score_column_matcher_prefers_first_regular_subject_and_preserves_original_name(self):
        self.assertTrue(
            hasattr(grade_logic, "find_first_matching_score_column"),
            "缺少 find_first_matching_score_column",
        )
        columns = ["班级", "姓名", "总成绩分数", "语文成绩", "数学分数"]

        self.assertEqual(grade_logic.find_first_matching_score_column(columns), "语文成绩")
        self.assertTrue(has_analyzable_columns(columns))
        self.assertEqual(find_first_matching_column(columns, NAME_COLUMN_ALIASES), "姓名")
        self.assertEqual(find_first_matching_column(columns, CLASS_COLUMN_ALIASES), "班级")

    def test_score_column_recognition_rejects_descriptive_columns(self):
        self.assertTrue(
            hasattr(grade_logic, "find_first_matching_score_column"),
            "缺少 find_first_matching_score_column",
        )
        descriptive_columns = ["班级", "姓名", "考号", "学号", "排名", "名次", "班级分数线", "年级平均分"]

        self.assertIsNone(grade_logic.find_first_matching_score_column(descriptive_columns))
        self.assertFalse(has_analyzable_columns(descriptive_columns))

    def test_score_options_exclude_structural_columns_from_real_exam_sheet(self):
        self.assertTrue(hasattr(grade_logic, "get_non_score_columns"))
        self.assertTrue(hasattr(grade_logic, "build_score_column_options"))
        columns = ["班级", "考号", "姓名", "考室", "语文", "数学", "排名"]

        self.assertEqual(
            grade_logic.get_non_score_columns(columns),
            ["班级", "考号", "姓名", "考室", "排名"],
        )
        self.assertEqual(
            grade_logic.build_score_column_options(columns),
            ["语文", "数学"],
        )

    def test_score_options_keep_custom_subjects_and_reject_all_identity_aliases(self):
        build_options = getattr(grade_logic, "build_score_column_options", None)
        self.assertIsNotNone(build_options, "缺少统一成绩候选构造函数")
        columns = [
            "姓名",
            "学生姓名",
            "班级",
            "行政班",
            "学号",
            "学生学号",
            "学生编号",
            "学籍号",
            "考号",
            "准考证号",
            "数学成绩",
            "音乐",
            "校本课程",
        ]

        self.assertEqual(
            build_options(columns),
            ["数学成绩", "音乐", "校本课程"],
        )

    def test_invalid_saved_score_column_falls_back_to_recognized_subject(self):
        build_options = getattr(grade_logic, "build_score_column_options", None)
        self.assertIsNotNone(build_options, "缺少统一成绩候选构造函数")
        columns = ["班级", "考号", "姓名", "考室", "语文", "数学", "排名"]
        score_options = build_options(columns)
        matched_score = grade_logic.find_first_matching_score_column(columns)

        self.assertEqual(
            grade_logic.resolve_column_selection(
                score_options,
                "考号",
                matched_score,
            ),
            "语文",
        )

    def test_header_detection_accepts_suffixed_score_columns(self):
        self.assertTrue(
            hasattr(grade_logic, "find_first_matching_score_column"),
            "缺少 find_first_matching_score_column",
        )
        raw_df = pd.DataFrame(
            [
                ["班级", "姓名", "语文分数", "数学分数", "总分分数"],
                ["1班", "张三", 110, 115, 225],
            ]
        )

        self.assertEqual(detect_header_row(raw_df), 0)
        self.assertTrue(has_analyzable_columns(raw_df.iloc[0].tolist()))
        self.assertEqual(
            grade_logic.find_first_matching_score_column(raw_df.iloc[0].tolist()),
            "语文分数",
        )

    def test_full_score_helpers_support_isolated_column_settings_and_total_notice(self):
        required_helpers = [
            "build_full_score_context_key",
            "get_column_full_score",
            "set_column_full_score",
            "suggest_full_score",
            "get_total_score_notice",
        ]
        for helper_name in required_helpers:
            self.assertTrue(hasattr(grade_logic, helper_name), f"缺少 {helper_name}")

        math_context = grade_logic.build_full_score_context_key(b"workbook-a", "成绩表")
        other_sheet_context = grade_logic.build_full_score_context_key(b"workbook-a", "第二学期")
        other_file_context = grade_logic.build_full_score_context_key(b"workbook-b", "成绩表")
        settings = {}

        self.assertEqual(grade_logic.get_column_full_score(settings, math_context, "数学"), 120.0)
        grade_logic.set_column_full_score(settings, math_context, "数学", 125)
        self.assertEqual(grade_logic.get_column_full_score(settings, math_context, "总分"), 800.0)
        grade_logic.set_column_full_score(settings, math_context, "总分", 850)

        self.assertEqual(grade_logic.get_column_full_score(settings, math_context, "数学"), 125.0)
        self.assertEqual(grade_logic.get_column_full_score(settings, math_context, "总分"), 850.0)
        self.assertEqual(grade_logic.get_column_full_score(settings, other_sheet_context, "数学"), 120.0)
        self.assertEqual(grade_logic.get_column_full_score(settings, other_file_context, "数学"), 120.0)
        self.assertNotEqual(math_context, other_sheet_context)
        self.assertNotEqual(math_context, other_file_context)
        self.assertEqual(grade_logic.suggest_full_score("未知成绩列"), 100.0)
        self.assertEqual(grade_logic.suggest_full_score("语文成绩"), 120.0)
        self.assertEqual(
            grade_logic.get_total_score_notice("期末总成绩"),
            "当前选择的是总分列，请确认总分满分，避免有效成绩被错误过滤。",
        )
        self.assertIsNone(grade_logic.get_total_score_notice("数学"))

    def test_unsafe_full_score_suggestion_requires_confirmation_and_is_not_adopted(self):
        self.assertTrue(hasattr(grade_logic, "get_full_score_suggestion"))
        for unsafe_value in (5.0, 1201.0):
            with self.subTest(unsafe_value=unsafe_value):
                settings = {}
                with patch("grade_logic.suggest_full_score", return_value=unsafe_value):
                    suggestion = grade_logic.get_full_score_suggestion("数学")
                    adopted = grade_logic.get_column_full_score(
                        settings,
                        "workbook:sheet",
                        "数学",
                    )

                self.assertEqual(suggestion.value, 100.0)
                self.assertEqual(suggestion.suggested_value, unsafe_value)
                self.assertTrue(suggestion.requires_confirmation)
                self.assertEqual(adopted, 100.0)

    def test_full_score_survives_name_field_change(self):
        settings = {"exam:sheet": {"数学": 120.0}}
        session_state = {"analysis_name_column": "姓名"}

        key, value = grade_logic.initialize_full_score_widget_state(
            session_state,
            settings,
            "exam:sheet",
            "数学",
        )
        session_state["analysis_name_column"] = "学生姓名"
        regenerated_key, regenerated_value = (
            grade_logic.initialize_full_score_widget_state(
                session_state,
                settings,
                "exam:sheet",
                "数学",
            )
        )

        self.assertEqual(regenerated_key, key)
        self.assertEqual(regenerated_value, value)
        self.assertEqual(regenerated_value, 120.0)

    def test_full_score_survives_class_field_change(self):
        settings = {"exam:sheet": {"数学": 120.0}}
        session_state = {"analysis_class_column": "班级"}

        key, value = grade_logic.initialize_full_score_widget_state(
            session_state,
            settings,
            "exam:sheet",
            "数学",
        )
        session_state["analysis_class_column"] = "行政班"
        regenerated_key, regenerated_value = (
            grade_logic.initialize_full_score_widget_state(
                session_state,
                settings,
                "exam:sheet",
                "数学",
            )
        )

        self.assertEqual(regenerated_key, key)
        self.assertEqual(regenerated_value, value)
        self.assertEqual(regenerated_value, 120.0)

    def test_full_score_widget_key_normalizes_column_and_restores_each_subject(self):
        settings = {"exam:sheet": {"数学": 120.0, "英语": 150.0}}
        session_state = {}

        math_key, math_value = grade_logic.initialize_full_score_widget_state(
            session_state,
            settings,
            "exam:sheet",
            " 数学\n",
        )
        english_key, english_value = grade_logic.initialize_full_score_widget_state(
            session_state,
            settings,
            "exam:sheet",
            "英语",
        )
        restored_math_key, restored_math_value = (
            grade_logic.initialize_full_score_widget_state(
                session_state,
                settings,
                "exam:sheet",
                "数学",
            )
        )

        self.assertEqual(math_key, "full_score::exam:sheet::数学")
        self.assertEqual(restored_math_key, math_key)
        self.assertNotEqual(english_key, math_key)
        self.assertEqual((math_value, english_value, restored_math_value), (120.0, 150.0, 120.0))

    def test_regenerated_widget_cannot_replace_saved_full_score_with_one(self):
        settings = {"exam:sheet": {"数学": 120.0}}
        session_state = {"full_score::exam:sheet::数学": 1.0}

        key, initialized_value = grade_logic.initialize_full_score_widget_state(
            session_state,
            settings,
            "exam:sheet",
            "数学",
        )
        effective_value = grade_logic.set_column_full_score_safely(
            settings,
            "exam:sheet",
            "数学",
            session_state[key],
        )

        self.assertEqual(session_state[key], 120.0)
        self.assertEqual(initialized_value, 120.0)
        self.assertEqual(effective_value, 120.0)
        self.assertEqual(settings["exam:sheet"]["数学"], 120.0)

    def test_user_input_one_keeps_existing_full_score(self):
        settings = {"exam:sheet": {"数学": 120.0}}

        effective_value = grade_logic.set_column_full_score_safely(
            settings,
            "exam:sheet",
            "数学",
            1.0,
        )

        self.assertEqual(effective_value, 120.0)
        self.assertEqual(settings["exam:sheet"]["数学"], 120.0)

    def test_analyze_scores_builds_summary_and_lists(self):
        result = analyze_scores({"Alice": 95, "Bob": 82, "Cindy": 76, "Dan": 58})

        self.assertEqual(result["student_count"], 4)
        self.assertEqual(result["highest_score"], 95)
        self.assertEqual(result["lowest_score"], 58)
        self.assertAlmostEqual(result["average_score"], 77.75)
        self.assertAlmostEqual(result["excellent_rate"], 25.0)
        self.assertAlmostEqual(result["pass_rate"], 75.0)
        self.assertEqual(result["excellent_students"], [["Alice", 95]])
        self.assertEqual(result["fail_students"], [["Dan", 58]])
        self.assertEqual(get_score_level(59.9), "不及格")
        self.assertEqual(get_score_level(90), "优秀")

    def test_export_score_result_returns_downloadable_workbook_bytes(self):
        result = analyze_scores({"Alice": 95, "Bob": 58})
        data = export_score_result_to_bytes(result)

        self.assertIsInstance(data, BytesIO)
        workbook = load_workbook(data)
        self.assertEqual(
            workbook.sheetnames,
            ["成绩明细", "基础统计", "优秀学生名单", "不及格学生名单"],
        )
        basic_values = {
            workbook["基础统计"].cell(row=row, column=1).value: workbook["基础统计"].cell(row=row, column=2).value
            for row in range(1, workbook["基础统计"].max_row + 1)
        }
        self.assertEqual(workbook["成绩明细"]["A1"].value, "名次")
        self.assertEqual(basic_values["总人数"], 2)
        self.assertEqual(workbook["优秀学生名单"]["A2"].value, "Alice")
        self.assertEqual(workbook["不及格学生名单"]["A2"].value, "Bob")

    def test_create_single_score_template_returns_styled_example_workbook(self):
        data = create_single_score_template()

        self.assertIsInstance(data, BytesIO)
        workbook = load_workbook(data)
        self.assertEqual(workbook.sheetnames, ["成绩录入"])

        sheet = workbook["成绩录入"]
        self.assertEqual([sheet["A1"].value, sheet["B1"].value], ["姓名", "分数"])
        self.assertEqual(
            [[sheet.cell(row=row, column=1).value, sheet.cell(row=row, column=2).value] for row in range(2, 6)],
            [["张三", 85], ["李四", 72], ["王五", 96], ["赵六", 58]],
        )
        self.assertTrue(sheet["A1"].font.bold)
        self.assertEqual(sheet["A1"].alignment.horizontal, "center")
        self.assertGreaterEqual(sheet.column_dimensions["A"].width, 6)

    def test_clean_column_name_removes_spaces_newlines_and_full_width_spaces(self):
        self.assertEqual(clean_column_name("　姓名 \n"), "姓名")
        self.assertEqual(clean_column_name("学生\r\n姓名"), "学生姓名")
        self.assertEqual(clean_column_name(123), "123")

    def test_find_first_matching_column_supports_score_aliases(self):
        columns = ["学生姓名", "数学"]

        self.assertEqual(find_first_matching_column(columns, ["姓名", "学生姓名"]), "学生姓名")
        self.assertEqual(find_first_matching_column(columns, ["分数", "成绩", "数学"]), "数学")
        self.assertIsNone(find_first_matching_column(columns, ["总分"]))

    def test_has_analyzable_columns_requires_name_and_score_or_subject(self):
        self.assertTrue(has_analyzable_columns(["班级", "姓名", "数学"]))
        self.assertTrue(has_analyzable_columns(["学生姓名", "成绩"]))
        self.assertFalse(has_analyzable_columns(["班级", "姓名", "备注"]))
        self.assertFalse(has_analyzable_columns(["班级", "数学", "备注"]))

    def test_detect_header_row_finds_school_report_header_on_second_row(self):
        raw_df = pd.DataFrame(
            [
                ["零陵区实验中学2026年上期成绩汇总", None, None, None, None, None],
                ["班级", "考号", "姓名", "考室", "语文", "数学"],
                ["七1班", "001", "张三", "A01", 108, 117],
                ["七1班", "002", "李四", "A01", 96, 98],
            ]
        )

        self.assertEqual(detect_header_row(raw_df), 1)
        df = build_dataframe_from_header(raw_df, 1)
        self.assertEqual(df.columns.tolist(), ["班级", "考号", "姓名", "考室", "语文", "数学"])
        self.assertEqual(df["数学"].max(), 117)

    def test_analyze_scores_uses_full_score_for_levels_and_rates(self):
        result = analyze_scores({"张三": 117, "李四": 98, "王五": 71}, full_score=120)

        self.assertEqual(result["highest_score"], 117)
        self.assertEqual(result["score_details"][0], ["张三", 117, "优秀"])
        self.assertEqual(result["score_details"][1], ["李四", 98, "良好"])
        self.assertEqual(result["score_details"][2], ["王五", 71, "不及格"])
        self.assertAlmostEqual(result["excellent_rate"], 100 / 3)
        self.assertAlmostEqual(result["pass_rate"], 200 / 3)
        self.assertEqual(get_score_level(108, full_score=120), "优秀")
        self.assertEqual(get_score_level(96, full_score=120), "良好")
        self.assertEqual(get_score_level(72, full_score=120), "及格")

    def test_analyze_scores_uses_custom_excellent_percent(self):
        result = analyze_scores({"张三": 85, "李四": 84, "王五": 59}, excellent_percent=85)

        self.assertEqual(result["excellent_count"], 1)
        self.assertEqual(result["fail_count"], 1)
        self.assertAlmostEqual(result["excellent_rate"], 100 / 3)
        self.assertEqual(result["excellent_students"], [["张三", 85]])
        self.assertEqual(result["fail_students"], [["王五", 59]])
        self.assertEqual(result["score_details"][0], ["张三", 85, "优秀"])
        self.assertEqual(result["excellent_percent"], 85)
        self.assertEqual(result["pass_percent"], 60)

    def test_excellent_percent_below_pass_line_is_clamped_to_sixty(self):
        result = analyze_scores({"张三": 61, "李四": 59}, excellent_percent=50)

        self.assertEqual(result["excellent_percent"], 60)
        self.assertEqual(result["excellent_count"], 1)
        self.assertEqual(result["fail_count"], 1)
        self.assertEqual(result["score_details"][0], ["张三", 61, "优秀"])

    def test_export_score_result_includes_current_scoring_settings(self):
        result = analyze_scores({"张三": 85, "李四": 59}, full_score=100, excellent_percent=85)
        data = export_score_result_to_bytes(result)

        workbook = load_workbook(data)
        basic_sheet = workbook["基础统计"]
        values = {basic_sheet.cell(row=row, column=1).value: basic_sheet.cell(row=row, column=2).value for row in range(1, basic_sheet.max_row + 1)}

        self.assertEqual(values["满分"], 100)
        self.assertEqual(values["优秀线"], "85%")
        self.assertEqual(values["及格线"], "60%")
        self.assertEqual(values["优秀人数"], 1)
        self.assertEqual(values["不及格人数"], 1)

    def test_class_aliases_and_options_support_numeric_and_text_classes(self):
        self.assertIn("班级", CLASS_COLUMN_ALIASES)
        self.assertIn("行政班", CLASS_COLUMN_ALIASES)

        class_values = pd.Series([2401.0, "2402班", 2401, None, " 2403 "])

        self.assertEqual(format_class_value(2401.0), "2401")
        self.assertEqual(format_class_value("2401班"), "2401班")
        self.assertEqual(build_class_options(class_values), ["全部班级", "2401", "2402班", "2403"])

    def test_single_class_mode_includes_all_students_and_preserves_valid_selection(self):
        class_values = pd.Series([2501, 2502, 2503, 2504, 2505, 2501])

        options = grade_logic.build_single_class_options(class_values)

        self.assertEqual(options, ["全部学生", "2501", "2502", "2503", "2504", "2505"])
        self.assertNotIn("全部班级", options)
        self.assertEqual(grade_logic.resolve_single_class_selection(options), "2501")
        self.assertEqual(grade_logic.resolve_single_class_selection(options, "全部学生"), "全部学生")
        self.assertEqual(grade_logic.resolve_single_class_selection(options, "2502"), "2502")
        self.assertEqual(grade_logic.resolve_single_class_selection(options, "全部班级"), "2501")
        self.assertEqual(grade_logic.resolve_single_class_selection(options, "不存在的班级"), "2501")
        self.assertEqual(grade_logic.resolve_single_class_selection([], None), "全部学生")

    def test_all_students_selection_survives_subject_and_sheet_changes(self):
        original_options = grade_logic.build_single_class_options(pd.Series([2501, 2502]))
        new_sheet_options = grade_logic.build_single_class_options(pd.Series([2601, 2602]))

        self.assertEqual(
            grade_logic.resolve_single_class_selection(original_options, "全部学生"),
            "全部学生",
        )
        self.assertEqual(
            grade_logic.resolve_single_class_selection(new_sheet_options, "全部学生"),
            "全部学生",
        )
        self.assertEqual(
            grade_logic.resolve_single_class_selection(new_sheet_options, "2502"),
            "2601",
        )

    def test_single_class_filter_strictly_isolates_2501_and_2502(self):
        dataframe = pd.DataFrame(
            {
                "班级": [2501, 2501, 2502, 2502],
                "姓名": ["甲", "乙", "丙", "丁"],
                "数学": [95, 58, 88, 76],
            }
        )

        class_2501 = grade_logic.filter_dataframe_by_class(dataframe, "班级", "2501")
        class_2502 = grade_logic.filter_dataframe_by_class(dataframe, "班级", "2502")
        all_students = grade_logic.filter_dataframe_by_class(dataframe, "班级", "全部学生")

        self.assertEqual(class_2501["姓名"].tolist(), ["甲", "乙"])
        self.assertEqual(class_2502["姓名"].tolist(), ["丙", "丁"])
        self.assertEqual(all_students["姓名"].tolist(), ["甲", "乙", "丙", "丁"])
        self.assertEqual(class_2501["班级"].apply(format_class_value).unique().tolist(), ["2501"])
        self.assertEqual(class_2502["班级"].apply(format_class_value).unique().tolist(), ["2502"])
        with self.assertRaises(ValueError):
            grade_logic.filter_dataframe_by_class(dataframe, "班级", "全部班级")

    def test_all_students_analysis_and_excel_export_use_the_full_filtered_scope(self):
        dataframe = pd.DataFrame(
            {
                "班级": [2501, 2501, 2502],
                "姓名": ["甲", "乙", "丙"],
                "数学": [95, 58, 88],
            }
        )
        analysis_df = grade_logic.filter_dataframe_by_class(dataframe, "班级", "全部学生")
        result = analyze_scores(
            dict(zip(analysis_df["姓名"], analysis_df["数学"])),
            current_class="全部学生",
            current_subject="数学",
        )

        workbook = load_workbook(export_score_result_to_bytes(result))
        exported_names = [
            workbook["成绩明细"].cell(row=row, column=2).value
            for row in range(2, workbook["成绩明细"].max_row + 1)
        ]

        self.assertEqual(len(analysis_df), 3)
        self.assertEqual(result["student_count"], 3)
        self.assertEqual(result["excellent_students"], [["甲", 95]])
        self.assertEqual(result["fail_students"], [["乙", 58]])
        self.assertEqual(exported_names, ["甲", "丙", "乙"])

    def test_missing_class_column_analyzes_all_students_without_filtering(self):
        dataframe = pd.DataFrame(
            {"姓名": ["甲", "乙", "丙"], "数学": [95, 58, 88]}
        )

        analysis_df = grade_logic.filter_dataframe_by_class(
            dataframe,
            None,
            "全部学生",
        )

        self.assertEqual(analysis_df.to_dict("records"), dataframe.to_dict("records"))
        self.assertIsNot(analysis_df, dataframe)

    def test_filtered_single_class_analysis_and_excel_export_only_include_current_class(self):
        dataframe = pd.DataFrame(
            {
                "班级": [2501, 2501, 2502],
                "姓名": ["甲", "乙", "丙"],
                "数学": [95, 58, 88],
            }
        )
        filtered = grade_logic.filter_dataframe_by_class(dataframe, "班级", "2501")
        result = analyze_scores(
            dict(zip(filtered["姓名"], filtered["数学"])),
            current_class="2501",
            current_subject="数学",
        )

        workbook = load_workbook(export_score_result_to_bytes(result))
        exported_names = [
            workbook["成绩明细"].cell(row=row, column=2).value
            for row in range(2, workbook["成绩明细"].max_row + 1)
        ]

        self.assertEqual(result["student_count"], 2)
        self.assertEqual(exported_names, ["甲", "乙"])
        self.assertNotIn("丙", exported_names)

    def test_column_selection_falls_back_when_previous_sheet_column_is_missing(self):
        new_sheet_columns = ["班级", "姓名", "数学", "英语"]

        self.assertEqual(
            grade_logic.resolve_column_selection(new_sheet_columns, "语文", "数学", fallback_index=1),
            "数学",
        )
        self.assertEqual(
            grade_logic.resolve_column_selection(new_sheet_columns, "英语", "数学", fallback_index=1),
            "英语",
        )

    def test_analyze_scores_and_export_include_current_class_and_subject(self):
        result = analyze_scores(
            {"张三": 117, "李四": 98},
            full_score=120,
            excellent_percent=90,
            current_class="2401",
            current_subject="数学",
        )
        data = export_score_result_to_bytes(result)

        workbook = load_workbook(data)
        basic_sheet = workbook["基础统计"]
        values = {basic_sheet.cell(row=row, column=1).value: basic_sheet.cell(row=row, column=2).value for row in range(1, basic_sheet.max_row + 1)}

        self.assertEqual(result["current_class"], "2401")
        self.assertEqual(result["current_subject"], "数学")
        self.assertEqual(values["当前班级"], "2401")
        self.assertEqual(values["当前分析科目"], "数学")
        self.assertEqual(values["满分"], 120)
        self.assertEqual(values["优秀线"], "90%")
        self.assertEqual(values["及格线"], "60%")

    def test_export_without_class_column_labels_scope_as_all_students(self):
        result = analyze_scores(
            {"张三": 90, "李四": 80},
            current_class="全部学生",
            current_subject="数学",
        )

        workbook = load_workbook(export_score_result_to_bytes(result))
        basic_sheet = workbook["基础统计"]
        values = {
            basic_sheet.cell(row=row, column=1).value: basic_sheet.cell(row=row, column=2).value
            for row in range(1, basic_sheet.max_row + 1)
        }

        self.assertEqual(values["分析对象"], "全部学生")
        self.assertNotIn("当前班级", values)


if __name__ == "__main__":
    unittest.main()
