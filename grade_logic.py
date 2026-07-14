import hashlib
from io import BytesIO

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side


NAME_COLUMN_ALIASES = ["姓名", "学生姓名", "名字", "姓名列"]
SUBJECT_COLUMN_ALIASES = [
    "数学",
    "语文",
    "英语",
    "物理",
    "化学",
    "道法",
    "历史",
    "地理",
    "生物",
    "政治",
    "体育",
    "信息技术",
]
SCORE_COLUMN_ALIASES = ["分数", "成绩", *SUBJECT_COLUMN_ALIASES, "总分"]
CLASS_COLUMN_ALIASES = ["班级", "班别", "行政班", "班级名称"]
TOTAL_SCORE_KEYWORDS = ("总分", "总成绩", "合计")
FULL_SCORE_NOTICE = "当前选择的是总分列，请确认总分满分，避免有效成绩被错误过滤。"


def clean_column_name(column_name):
    text = str(column_name)
    text = text.replace("\r", "").replace("\n", "").replace("\u3000", "")
    return text.strip()


def clean_dataframe_columns(dataframe):
    cleaned = dataframe.copy()
    cleaned.columns = [clean_column_name(column) for column in cleaned.columns]
    return cleaned


def is_total_score_column(column_name):
    cleaned_name = clean_column_name(column_name)
    return any(keyword in cleaned_name for keyword in TOTAL_SCORE_KEYWORDS)


def suggest_full_score(column_name):
    cleaned_name = clean_column_name(column_name)
    if is_total_score_column(cleaned_name):
        return 800.0
    if any(subject in cleaned_name for subject in ("语文", "数学", "英语")):
        return 120.0
    return 100.0


def get_total_score_notice(column_name):
    if is_total_score_column(column_name):
        return FULL_SCORE_NOTICE
    return None


def build_full_score_context_key(file_content, sheet_name):
    file_fingerprint = hashlib.sha256(file_content).hexdigest()
    return f"{file_fingerprint}:{clean_column_name(sheet_name)}"


def get_column_full_score(settings, context_key, column_name):
    context_settings = settings.setdefault(context_key, {})
    cleaned_name = clean_column_name(column_name)
    if cleaned_name not in context_settings:
        context_settings[cleaned_name] = suggest_full_score(cleaned_name)
    return float(context_settings[cleaned_name])


def set_column_full_score(settings, context_key, column_name, full_score):
    context_settings = settings.setdefault(context_key, {})
    context_settings[clean_column_name(column_name)] = float(full_score)


def find_first_matching_column(columns, aliases):
    for alias in aliases:
        for column in columns:
            if clean_column_name(column) == alias:
                return column
    return None


def has_analyzable_columns(columns):
    return (
        find_first_matching_column(columns, NAME_COLUMN_ALIASES) is not None
        and find_first_matching_column(columns, SCORE_COLUMN_ALIASES) is not None
    )


def format_class_value(value):
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except NameError:
        pass
    except TypeError:
        pass

    if isinstance(value, float) and value.is_integer():
        return str(int(value))

    text = str(value).strip()
    if text.endswith(".0"):
        numeric_text = text[:-2]
        if numeric_text.isdigit():
            return numeric_text
    return text


def build_class_options(class_values):
    seen = set()
    options = ["全部班级"]
    for value in class_values:
        formatted = format_class_value(value)
        if formatted and formatted not in seen:
            seen.add(formatted)
            options.append(formatted)
    return options


def detect_header_row(raw_dataframe, max_scan_rows=15):
    scan_rows = min(max_scan_rows, len(raw_dataframe))
    for row_index in range(scan_rows):
        columns = [clean_column_name(value) for value in raw_dataframe.iloc[row_index].tolist()]
        if (
            find_first_matching_column(columns, NAME_COLUMN_ALIASES) is not None
            and find_first_matching_column(columns, SCORE_COLUMN_ALIASES) is not None
        ):
            return row_index
    return None


def build_dataframe_from_header(raw_dataframe, header_row_index):
    columns = [clean_column_name(value) for value in raw_dataframe.iloc[header_row_index].tolist()]
    data = raw_dataframe.iloc[header_row_index + 1 :].copy()
    data.columns = columns
    data = data.dropna(how="all")
    data = data.loc[:, [column != "" and column.lower() != "nan" for column in data.columns]]
    return data.reset_index(drop=True)


def normalize_excellent_percent(excellent_percent):
    return max(float(excellent_percent), 60)


def format_percent(percent):
    if float(percent).is_integer():
        return str(int(percent)) + "%"
    return str(round(percent, 1)).rstrip("0").rstrip(".") + "%"


def get_score_level(score, full_score=100, excellent_percent=90):
    excellent_percent = normalize_excellent_percent(excellent_percent)
    excellent_line = full_score * excellent_percent / 100
    good_line = full_score * 0.8
    pass_line = full_score * 0.6

    if score >= excellent_line:
        return "优秀"
    if score >= good_line:
        return "良好"
    if score >= pass_line:
        return "及格"
    return "不及格"


def analyze_scores(student_scores, full_score=100, excellent_percent=90, current_class="全部班级", current_subject=""):
    excellent_percent = normalize_excellent_percent(excellent_percent)
    score_details = []
    excellent_students = []
    fail_students = []
    total_score = 0
    excellent_count = 0
    good_count = 0
    pass_count = 0
    fail_count = 0
    highest_score = None
    lowest_score = None

    for name, score in student_scores.items():
        total_score += score

        if highest_score is None or score > highest_score:
            highest_score = score
        if lowest_score is None or score < lowest_score:
            lowest_score = score

        level = get_score_level(score, full_score, excellent_percent)
        if level == "优秀":
            excellent_count += 1
            excellent_students.append([name, score])
        elif level == "良好":
            good_count += 1
        elif level == "及格":
            pass_count += 1
        else:
            fail_count += 1
            fail_students.append([name, score])

        score_details.append([name, score, level])

    student_count = len(student_scores)
    if student_count > 0:
        average_score = total_score / student_count
        excellent_rate = excellent_count / student_count * 100
        real_pass_count = excellent_count + good_count + pass_count
        pass_rate = real_pass_count / student_count * 100
    else:
        average_score = 0
        excellent_rate = 0
        pass_rate = 0

    return {
        "score_details": sorted(score_details, key=lambda detail: detail[1], reverse=True),
        "student_count": student_count,
        "average_score": average_score,
        "highest_score": highest_score,
        "lowest_score": lowest_score,
        "excellent_count": excellent_count,
        "good_count": good_count,
        "pass_count": pass_count,
        "fail_count": fail_count,
        "excellent_rate": excellent_rate,
        "pass_rate": pass_rate,
        "excellent_students": sorted(excellent_students, key=lambda detail: detail[1], reverse=True),
        "fail_students": sorted(fail_students, key=lambda detail: detail[1]),
        "full_score": full_score,
        "excellent_percent": excellent_percent,
        "good_percent": 80,
        "pass_percent": 60,
        "current_class": current_class,
        "current_subject": current_subject,
    }


def beautify_sheet(sheet):
    sheet.freeze_panes = "A2"
    thin_side = Side(style="thin")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
            if cell.row == 1:
                cell.font = Font(bold=True)

    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        sheet.column_dimensions[column_letter].width = max_length + 4


def create_single_score_template():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "成绩录入"
    sheet.append(["姓名", "分数"])

    example_rows = [
        ["张三", 85],
        ["李四", 72],
        ["王五", 96],
        ["赵六", 58],
    ]
    for row in example_rows:
        sheet.append(row)

    beautify_sheet(sheet)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def export_score_result_to_bytes(analysis_result):
    workbook = Workbook()

    detail_sheet = workbook.active
    detail_sheet.title = "成绩明细"
    detail_sheet.append(["名次", "姓名", "分数", "等级"])
    for rank, detail in enumerate(analysis_result["score_details"], start=1):
        name, score, level = detail
        detail_sheet.append([rank, name, score, level])

    basic_sheet = workbook.create_sheet("基础统计")
    basic_sheet.append(["统计项目", "统计结果"])
    basic_sheet.append(["当前班级", analysis_result.get("current_class", "全部班级")])
    basic_sheet.append(["当前分析科目", analysis_result.get("current_subject", "")])
    basic_sheet.append(["满分", analysis_result["full_score"]])
    basic_sheet.append(["优秀线", format_percent(analysis_result["excellent_percent"])])
    basic_sheet.append(["良好线", format_percent(analysis_result["good_percent"])])
    basic_sheet.append(["及格线", format_percent(analysis_result["pass_percent"])])
    basic_sheet.append(["总人数", analysis_result["student_count"]])
    basic_sheet.append(["平均分", round(analysis_result["average_score"], 1)])
    basic_sheet.append(["最高分", analysis_result["highest_score"]])
    basic_sheet.append(["最低分", analysis_result["lowest_score"]])
    basic_sheet.append(["优秀人数", analysis_result["excellent_count"]])
    basic_sheet.append(["良好人数", analysis_result["good_count"]])
    basic_sheet.append(["及格人数", analysis_result["pass_count"]])
    basic_sheet.append(["不及格人数", analysis_result["fail_count"]])
    basic_sheet.append(["优秀率", str(round(analysis_result["excellent_rate"], 1)) + "%"])
    basic_sheet.append(["及格率", str(round(analysis_result["pass_rate"], 1)) + "%"])

    excellent_sheet = workbook.create_sheet("优秀学生名单")
    excellent_sheet.append(["姓名", "分数"])
    if analysis_result["excellent_students"]:
        for student in analysis_result["excellent_students"]:
            excellent_sheet.append([student[0], student[1]])
    else:
        excellent_sheet.append(["本次没有优秀学生。", ""])

    fail_sheet = workbook.create_sheet("不及格学生名单")
    fail_sheet.append(["姓名", "分数"])
    if analysis_result["fail_students"]:
        for student in analysis_result["fail_students"]:
            fail_sheet.append([student[0], student[1]])
    else:
        fail_sheet.append(["本次没有不及格学生。", ""])

    for sheet in workbook.worksheets:
        beautify_sheet(sheet)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output
